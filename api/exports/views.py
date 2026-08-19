import csv
import io
import unicodedata
from datetime import date, datetime
from io import BytesIO

from django.db import transaction as db_transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from transactions.categorization import categorize_free_text
from transactions.models import Account, Category, Transaction


def _month_bounds(month_str: str | None):
    if month_str:
        year, month = (int(p) for p in month_str.split("-"))
    else:
        today = date.today()
        year, month = today.year, today.month
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return start, end


def _rows(user, month_str):
    start, end = _month_bounds(month_str)
    qs = (
        Transaction.objects.filter(user=user, date__gte=start, date__lt=end)
        .select_related("category", "account")
        .order_by("date")
    )
    header = ["data", "descrição", "categoria", "valor", "origem", "conta"]
    yield header
    for tx in qs:
        yield [
            tx.date.isoformat(),
            tx.description,
            tx.category.name_pt,
            f"{tx.amount_cents / 100:.2f}",
            tx.origin,
            tx.account.name,
        ]


class ExportCsvView(APIView):
    def get(self, request):
        month = request.query_params.get("month")
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="finez-{month or "atual"}.csv"'
        writer = csv.writer(response)
        for row in _rows(request.user, month):
            writer.writerow(row)
        return response


class ExportXlsxView(APIView):
    def get(self, request):
        month = request.query_params.get("month")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Extrato"
        for row in _rows(request.user, month):
            sheet.append(row)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="finez-{month or "atual"}.xlsx"'
        return response


# --- Importação (seção 10 do pedido) -----------------------------------
# Aceita a mesma forma do nosso próprio export (colunas data/descrição/valor/
# categoria) e também planilhas de terceiros com cabeçalhos equivalentes.
_HEADER_ALIASES = {
    "data": "date",
    "date": "date",
    "descricao": "description",
    "descrição": "description",
    "description": "description",
    "valor": "amount",
    "amount": "amount",
    "categoria": "category",
    "category": "category",
}
_DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_header(raw: str) -> str | None:
    key = _strip_accents(str(raw).strip().lower())
    return _HEADER_ALIASES.get(key)


def _parse_date(raw) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount_cents(raw) -> int | None:
    if isinstance(raw, (int, float)):
        return round(raw * 100)
    text = str(raw).strip().replace("R$", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        # separador decimal é o último símbolo encontrado (1.234,56 ou 1,234.56)
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return round(float(text) * 100)
    except ValueError:
        return None


def _read_rows(uploaded_file):
    """Gera (linha_num, dict de células) a partir de um arquivo .csv ou .xlsx."""
    name = (uploaded_file.name or "").lower()
    if name.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return
        columns = [_normalize_header(cell) if cell is not None else None for cell in header]
        for line_num, row in enumerate(rows_iter, start=2):
            yield line_num, dict(zip(columns, row))
    else:
        text = io.TextIOWrapper(uploaded_file, encoding="utf-8-sig")
        reader = csv.reader(text)
        header = next(reader, None)
        if header is None:
            return
        columns = [_normalize_header(cell) for cell in header]
        for line_num, row in enumerate(reader, start=2):
            yield line_num, dict(zip(columns, row))


class ImportTransactionsView(APIView):
    """POST /api/imports/transactions — multipart, campo `file` (.csv ou .xlsx).

    Colunas esperadas: data, descrição, valor (sinal negativo = saída, positivo
    = entrada — mesma convenção do nosso export), categoria (opcional).
    """

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "envie um arquivo no campo 'file'."}, status=400)
        if not uploaded_file.name.lower().endswith((".csv", ".xlsx")):
            return Response({"detail": "formato não suportado — use .csv ou .xlsx."}, status=400)

        categories_by_name = {
            _strip_accents(c.name_pt.lower()): c for c in Category.objects.all()
        }
        categories_by_slug = {c.slug: c for c in Category.objects.all()}
        other_category = categories_by_slug[Category.Slug.OTHER]

        manual_account, _ = Account.objects.get_or_create(
            user=request.user,
            type=Account.Type.MANUAL,
            defaults={"name": "Lançamentos manuais"},
        )

        imported = 0
        errors = []
        to_create = []

        try:
            rows = list(_read_rows(uploaded_file))
        except Exception:
            return Response({"detail": "não foi possível ler o arquivo."}, status=400)

        for line_num, row in rows:
            row = {k: v for k, v in row.items() if k}
            raw_date, raw_desc, raw_amount = row.get("date"), row.get("description"), row.get("amount")
            if raw_date is None or raw_amount is None:
                errors.append(f"linha {line_num}: faltam colunas obrigatórias (data/valor)")
                continue

            parsed_date = _parse_date(raw_date)
            amount_cents = _parse_amount_cents(raw_amount)
            if parsed_date is None or amount_cents is None:
                errors.append(f"linha {line_num}: data ou valor inválido")
                continue

            description = str(raw_desc).strip() if raw_desc else "Importado"
            raw_category = str(row.get("category") or "").strip()
            category = categories_by_name.get(_strip_accents(raw_category.lower()))
            if category is None and raw_category:
                category = categories_by_slug.get(_strip_accents(raw_category.lower()).replace(" ", "_"))
            if category is None:
                guessed_category, _ = categorize_free_text(request.user, description)
                category = guessed_category or other_category

            to_create.append(
                Transaction(
                    user=request.user,
                    account=manual_account,
                    amount_cents=amount_cents,
                    description=description,
                    date=parsed_date,
                    category=category,
                    category_source=Transaction.CategorySource.USER,
                    origin=Transaction.Origin.WEB,
                )
            )
            imported += 1

        if to_create:
            with db_transaction.atomic():
                Transaction.objects.bulk_create(to_create)

        return Response({"imported": imported, "errors": errors})
