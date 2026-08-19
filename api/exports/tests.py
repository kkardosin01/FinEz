import io
from datetime import date

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from transactions.models import Account, Category, Transaction

pytestmark = pytest.mark.django_db


@pytest.fixture
def user():
    return User.objects.create_user(email="teste@finez.app", password="senha-forte-123")


@pytest.fixture
def account(user):
    return Account.objects.create(user=user, type=Account.Type.MANUAL, name="Manual")


@pytest.fixture
def categories():
    return {
        slug: Category.objects.create(slug=slug, name_pt=name, color_light="#000000", color_dark="#ffffff")
        for slug, name in [
            (Category.Slug.FOOD, "Alimentação"),
            (Category.Slug.GROCERIES, "Mercado"),
            (Category.Slug.OTHER, "Outros"),
        ]
    }


@pytest.fixture
def client(user):
    api_client = APIClient()
    api_client.force_authenticate(user=user)
    return api_client


# --- ExportCsvView -----------------------------------------------------------


def test_export_csv_contains_only_own_transactions_for_the_month(client, user, account, categories):
    other_user = User.objects.create_user(email="outra@finez.app", password="senha-forte-123")
    other_account = Account.objects.create(user=other_user, type=Account.Type.MANUAL, name="Manual")

    Transaction.objects.create(
        user=user, account=account, category=categories[Category.Slug.FOOD], amount_cents=-1500,
        description="lanche", date=date(2026, 8, 5),
        category_source=Transaction.CategorySource.USER, origin=Transaction.Origin.WEB,
    )
    Transaction.objects.create(
        user=user, account=account, category=categories[Category.Slug.FOOD], amount_cents=-500,
        description="lanche mes anterior", date=date(2026, 7, 20),
        category_source=Transaction.CategorySource.USER, origin=Transaction.Origin.WEB,
    )
    Transaction.objects.create(
        user=other_user, account=other_account, category=categories[Category.Slug.FOOD], amount_cents=-999,
        description="nao deveria aparecer", date=date(2026, 8, 5),
        category_source=Transaction.CategorySource.USER, origin=Transaction.Origin.WEB,
    )

    response = client.get("/api/exports/csv?month=2026-08")
    assert response.status_code == 200
    assert response["Content-Type"].startswith("text/csv")
    body = response.content.decode("utf-8")
    assert "lanche" in body
    assert "lanche mes anterior" not in body
    assert "nao deveria aparecer" not in body


def test_export_csv_requires_authentication():
    response = APIClient().get("/api/exports/csv")
    assert response.status_code in (401, 403)


# --- ExportXlsxView ------------------------------------------------------------


def test_export_xlsx_produces_valid_workbook_with_rows(client, user, account, categories):
    Transaction.objects.create(
        user=user, account=account, category=categories[Category.Slug.GROCERIES], amount_cents=-2000,
        description="mercado", date=date(2026, 8, 10),
        category_source=Transaction.CategorySource.USER, origin=Transaction.Origin.WEB,
    )

    response = client.get("/api/exports/xlsx?month=2026-08")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "data"
    assert any(row[1] == "mercado" for row in rows[1:])


# --- ImportTransactionsView ----------------------------------------------------


def _csv_file(content: str, name="import.csv"):
    return SimpleUploadedFile(name, content.encode("utf-8"), content_type="text/csv")


def test_import_csv_creates_transactions_with_matched_category(client, user, categories):
    csv_content = "data,descrição,valor,categoria\n2026-08-05,mercado da esquina,-45.90,Mercado\n"
    response = client.post("/api/imports/transactions", {"file": _csv_file(csv_content)}, format="multipart")

    assert response.status_code == 200
    assert response.data == {"imported": 1, "errors": []}
    transaction = Transaction.objects.get(user=user)
    assert transaction.amount_cents == -4590
    assert transaction.description == "mercado da esquina"
    assert transaction.category.slug == Category.Slug.GROCERIES
    assert transaction.origin == Transaction.Origin.WEB


def test_import_csv_without_category_falls_back_to_free_text_guess_or_other(client, user, categories):
    csv_content = "data,descrição,valor\n2026-08-05,despesa aleatoria,-10.00\n"
    response = client.post("/api/imports/transactions", {"file": _csv_file(csv_content)}, format="multipart")

    assert response.status_code == 200
    transaction = Transaction.objects.get(user=user)
    assert transaction.category.slug == Category.Slug.OTHER


def test_import_csv_with_brl_formatted_amount(client, user, categories):
    csv_content = 'data,descrição,valor\n05/08/2026,compra,"R$ 1.234,56"\n'
    response = client.post("/api/imports/transactions", {"file": _csv_file(csv_content)}, format="multipart")

    assert response.status_code == 200
    transaction = Transaction.objects.get(user=user)
    assert transaction.amount_cents == 123456
    assert transaction.date == date(2026, 8, 5)


def test_import_csv_reports_errors_for_invalid_rows_without_failing_whole_batch(client, user, categories):
    csv_content = (
        "data,descrição,valor\n"
        "2026-08-05,linha valida,-10.00\n"
        "data-invalida,linha invalida,-5.00\n"
        ",faltando data,-5.00\n"
    )
    response = client.post("/api/imports/transactions", {"file": _csv_file(csv_content)}, format="multipart")

    assert response.status_code == 200
    assert response.data["imported"] == 1
    assert len(response.data["errors"]) == 2
    assert Transaction.objects.filter(user=user).count() == 1


def test_import_without_file_is_rejected(client):
    response = client.post("/api/imports/transactions", {}, format="multipart")
    assert response.status_code == 400


def test_import_with_unsupported_extension_is_rejected(client):
    bad_file = SimpleUploadedFile("dados.txt", b"data,valor\n2026-08-05,-10\n", content_type="text/plain")
    response = client.post("/api/imports/transactions", {"file": bad_file}, format="multipart")
    assert response.status_code == 400


def test_import_isolates_transactions_by_user(client, user, categories):
    csv_content = "data,descrição,valor\n2026-08-05,despesa,-10.00\n"
    client.post("/api/imports/transactions", {"file": _csv_file(csv_content)}, format="multipart")

    other_user = User.objects.create_user(email="outra@finez.app", password="senha-forte-123")
    assert Transaction.objects.filter(user=user).count() == 1
    assert Transaction.objects.filter(user=other_user).count() == 0
